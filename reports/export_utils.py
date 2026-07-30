import csv
from django.http import HttpResponse
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def _get_dynamic_row_data(response, questions, is_public):
    row_data = []
    
    if is_public:
        ans_map = {ans.question_id: ans.answer for ans in response.publicformanswer_set.all()}
    else:
        ans_map = {ans.question_id: ans.answer for ans in response.formanswer_set.all()}

    for q in questions:
        ans = ans_map.get(q.id, "")
        if not ans or str(ans).strip() == "":
            ans = "N/A"
        row_data.append(ans)
        
    return row_data

def generate_dynamic_excel(form, queryset, questions, filename):
    is_public = form.access_type == 'public'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    headers = [q.question for q in questions]
    ws.append(headers)

    for response in queryset:
        row_data = _get_dynamic_row_data(response, questions, is_public)
        
        # Append to worksheet ensuring strings are saved explicitly for large numbers
        ws.append(row_data)
        
        # Format the recently appended row to explicitly be TEXT to prevent scientific notation
        for cell in ws[ws.max_row]:
            cell.number_format = '@'

    response_file = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response_file["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response_file)
    return response_file

def generate_dynamic_csv(form, queryset, questions, filename):
    is_public = form.access_type == 'public'
    
    response_file = HttpResponse(content_type='text/csv')
    response_file['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response_file)
    
    headers = [q.question for q in questions]
    writer.writerow(headers)
    
    for response in queryset:
        row_data = _get_dynamic_row_data(response, questions, is_public)
        writer.writerow(row_data)
        
    return response_file

def generate_dynamic_pdf(form, queryset, questions, filename):
    is_public = form.access_type == 'public'
    
    response_file = HttpResponse(content_type='application/pdf')
    response_file['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    # Use landscape for tabular reports
    doc = SimpleDocTemplate(response_file, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    
    story.append(Paragraph(f"{form.title} - Responses Report", title_style))
    
    # Create Table Data
    headers = [Paragraph(f"<b>{q.question}</b>", styles['Normal']) for q in questions]
    table_data = [headers]
    
    for response in queryset:
        row_data = _get_dynamic_row_data(response, questions, is_public)
        # Convert strings to Paragraphs for word wrapping in PDF
        para_row = [Paragraph(str(cell), styles['Normal']) for cell in row_data]
        table_data.append(para_row)
        
    if len(table_data) > 1:
        # Calculate dynamic widths based on number of columns
        num_cols = len(headers)
        col_width = (landscape(letter)[0] - 60) / max(1, num_cols)
        
        t = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No responses found.", styles['Normal']))
        
    doc.build(story)
    return response_file
